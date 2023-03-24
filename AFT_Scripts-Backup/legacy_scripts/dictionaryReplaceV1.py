# Reading and writing in excel can be done by single module
import openpyxl
from openpyxl.utils.cell import get_column_letter
  
workbook = openpyxl.load_workbook('raw1.xlsx')
dict = openpyxl.load_workbook('dict1.xlsx')

workbook.sheetnames
worksheet = workbook["Sheet1"]

dict.sheetnames
dict = dict["Sheet1"]
  
# Number of rows
number_of_columns = worksheet.max_column
  

replacementTextKeyPairs = {'Heading1': 'FirstName', 'Heading2': 'LastName','Heading3': 'DateOfBirth'}
  
# Iterate over the columns in the first row, search
# for the text and replace
for i in range(number_of_columns):
      
    cellValue = str(worksheet[get_column_letter(i+1)+str(1)].value)
      
    for key in replacementTextKeyPairs.keys():
          
        if str(cellValue) == key:
            newCellValue = replacementTextKeyPairs.get(key)
            worksheet[get_column_letter(i+1)+str(1)] = str(newCellValue)

workbook.save('output1.xlsx')